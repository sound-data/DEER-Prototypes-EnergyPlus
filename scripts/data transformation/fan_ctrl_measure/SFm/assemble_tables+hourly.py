#%%
import pandas as pd
import numpy as np
import openpyxl as op
import os
# %%
#patches merged cell format lost bug when reading/writing with openpyxl
#source: https://stackoverflow.com/questions/38734044/losing-merged-cells-border-while-editing-excel-file-with-openpyxl

from itertools import product
import types
import openpyxl
from openpyxl import worksheet
from openpyxl.utils import range_boundaries


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self._merged_cells:
            self._merged_cells.append(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.worksheet.merge_cells = merge_cells
patch_worksheet()
# %%
#copy and paste function, credit: https://yagisanatode.com/2017/11/18/copy-and-paste-ranges-in-excel-with-openpyxl-and-python-3/
#Define function for copying cells
#inputs: start cell, end cell, sheet we want to copy from

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #loops through selected rows
    for i in range(startRow, endRow+1, 1):
        rowSelected = []
        #appends the row to a rowSelected list
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the rowSelected list and nests inside the rangeSelected
        
        rangeSelected.append(rowSelected)
    
    return rangeSelected


#Define function for pasting cells
#inputs: start cell, end cell, receiving sheet, copiedData)

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow+1, 1):
        countCol = 0
        for j in range(startCol, endCol+1, 1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol = countCol+1
        countRow = countRow + 1

# %%
#Pick only ONE of the set up paths

#%%
###Set up
##2-story path, SFm


enduses = 'end_uses_combined_2s.xlsx'
fan_split1 = 'fan_split1_combined_2s.xlsx'
fan_split2 = 'fan_split2_combined_2s.xlsx'
fan_table = 'fan_table_combined_2s.xlsx'
airloophvac = 'airloophvac_2s.xlsx'
combined_8760 = 'combined_2_story_8760s.xlsx'

# %%
##1-story path
#1-story path, SFm


enduses = 'end_uses_combined_1s.xlsx'
fan_split1 = 'fan_split1_combined_1s.xlsx'
fan_split2 = 'fan_split2_combined_1s.xlsx'
fan_table = 'fan_table_combined_1s.xlsx'
airloophvac = 'airloophvac_1s.xlsx'
combined_8760 = 'combined_1_story_8760s.xlsx'


# %%
####Read master template before for loop
#set up destination workbook (Master Template)
print('reading Template...')
wb = op.load_workbook('Fan_Control_Savings_Template_v3.xlsx')

#set up CZ iterator and run for loop
cz_list = ['CZ01','CZ02','CZ03','CZ04','CZ05','CZ06','CZ07','CZ08','CZ09','CZ10','CZ11','CZ12','CZ13','CZ14','CZ15','CZ16']

wb_eu = op.load_workbook(enduses)
wb_fan = op.load_workbook(fan_table)
wb_fs1 = op.load_workbook(fan_split1)
wb_fs2 = op.load_workbook(fan_split2)
wb_airloop = op.load_workbook(airloophvac)
for cz in cz_list:
    #read each table from 5 workbook, from respective CZ tab
    print(f'copying data from {cz}...')
    ws_eu = wb_eu[cz]
    enduse_extracted = copyRange(1,1,14,18,ws_eu)

    ws_fan = wb_fan[cz]
    fantable_extracted = copyRange(1,1,12,4,ws_fan)

    ws_fs1 = wb_fs1[cz]
    fs1_extracted = copyRange(1,1,6,18,ws_fs1)

    ws_fs2 = wb_fs2[cz]
    fs2_extracted = copyRange(1,1,6,18,ws_fs2)

    ws_airloop = wb_airloop[cz]
    airloop_extracted = copyRange(1,1,10,4,ws_airloop)

    print(f'pasting data to {cz}_Tables...')
    #setup destination worksheet
    ws_destination = wb[cz + '_Tables']
    #Paste values
    #enduse table
    pasteRange(1,1,14,18,ws_destination,enduse_extracted)
    #fan table
    pasteRange(1,22,12,25,ws_destination,fantable_extracted)
    #fan split 1
    pasteRange(1,29,6,46,ws_destination,fs1_extracted)
    #fan split 2
    pasteRange(8,29,13,46,ws_destination,fs2_extracted)
    #airloop hvac
    pasteRange(1,50,10,53,ws_destination,airloop_extracted)
print('parsed all html tables, doing hourly next')

# %%
####Start of for loop over 16 CZ to transfer 8760 data

for cz in cz_list:
    #read respective CZ hourly data
    print(f'extracting {cz} hourly 8760 data:')
    df_hourly = pd.read_excel(combined_8760, sheet_name=cz)
    #strip off design day hourly data to only get 8760 full year
    record_count_diff = len(df_hourly) - 8760
    df_hrly_cleaned = df_hourly.iloc[record_count_diff:].reset_index(drop=True)
    #isolate column that contains 'Part Load Ratio' (should be for both compressor and fan)
    #If want to look at any other hourly column, add them here
    df_filtered = df_hrly_cleaned.filter(like='Part Load Ratio')
    #read in rated fan power from fan table
    df_fantable = pd.read_excel(fan_table, sheet_name=cz, skiprows=1)
    rated_fanpower_list = df_fantable['Rated Power Per Max Air Flow Rate [W-s/m3]']
    rated_elec = rated_fanpower_list[0]/1000
    rated_elec2 = rated_fanpower_list[1]/1000

    print('caluclating fan cooling energy from PLR...')
    #calculate fan cooling energy if compressor is on for each fan
    df = df_filtered.copy()
    df['fan1_cooling_energy(kWh)'] = np.where(df.iloc[:,1] > 0, (df.iloc[:,0])*rated_elec, 0)
    df['fan2_cooling_energy(kWh)'] = np.where(df.iloc[:,3] > 0, (df.iloc[:,2])*rated_elec2, 0)
    #export to a temp excel for openpyxl copy and paste
    df.to_excel('temp.xlsx', index=False)
    wb_temp = op.load_workbook('temp.xlsx')
    ws_temp = wb_temp['Sheet1']

    print('copying and pasting value to template...')
    #copy all hourly data
    hourly_data_extracted = copyRange(1,1,6,8761,ws_temp)

    #setup paste destination
    ws_destination = wb[cz + '_hrly_var']
    #paste into destination cells
    pasteRange(2,1,7,8761, ws_destination, hourly_data_extracted)
print('go to next step.')

# %%

#populate bldg conditioned area for 1 story

#Normunit for AC measure: Cap-Ton - 9/1/22: changed to Area-ft2-BA
#using htm beautiful soup extraction
from bs4 import BeautifulSoup
import re
#9/1/2022 actual SEER rated baseline output path, SFm

#%%
path = "../../../../Analysis"
#grab path
file_list = os.listdir(path)
#list comprehension to extract only folder names
vint_folder_list = [i for i in file_list if 'SFm_SEER Rated AC_HP_1' in i]

#%%
# path = "High SEER baseline"
# file_list = os.listdir(path)
# vint_folder_list = [i for i in file_list if 'SFm' in i]

#%%
overall_path = path

cz_list = ['CZ01','CZ02','CZ03','CZ04','CZ05','CZ06','CZ07','CZ08','CZ09','CZ10','CZ11','CZ12','CZ13','CZ14','CZ15','CZ16']

# %%

#Parse output data from html output file of each CZ (measure case)
#1-story (vint_folder_list1)
conditioned_area_list = []
for vint_folder in vint_folder_list:
    print(f"parsing CZ specific building area from {vint_folder}")
    
    full_path = overall_path + "/" + vint_folder + "/" + "runs"
    cz_folders = [cz for cz in os.listdir(full_path) if 'CZ' in cz]
   
    for cz in cz_folders:
        print(f"parsing {cz} data...")
        
        subpath = full_path + "/" + f"{cz}" + "/" + "SFm&1&rDXGF&Ex&dxAC_equip/dxAC-Res-SEER-13.0/instance-tbl.htm"
        
        page=open(subpath)
        soup = BeautifulSoup(page.read(), features="lxml")
        
        raw_list = list(soup.find_all('tr'))
        extracted_info = []
        for i in range(0, len(raw_list)):
            if ">Net Conditioned Building Area<" in str(raw_list[i]):
                net_conditioned_area = float(str(raw_list[i])[-17:-11])
            if ">Total Building Area<" in str(raw_list[i]):
                total_area = float(str(raw_list[i])[-17:-11])
        #originally, index -21:-12
        print(f"{cz},conditioned_area = {net_conditioned_area}(m2), total_area = {total_area}(m2)")
        conditioned_area_list.append(net_conditioned_area)
# %%
conditioned_area_list_2s = []
for vint_folder in vint_folder_list:
    print(f"parsing CZ specific building area from {vint_folder}")
    
    full_path = overall_path + "/" + vint_folder + "/" + "runs"
    cz_folders = [cz for cz in os.listdir(full_path) if 'CZ' in cz]
   
    for cz in cz_folders:
        print(f"parsing {cz} data...")
        
        subpath = full_path + "/" + f"{cz}" + "/" + "SFm&2&rDXGF&Ex&dxAC_equip/dxAC-Res-SEER-13.0/instance-tbl.htm"
        
        page=open(subpath)
        soup = BeautifulSoup(page.read(), features="lxml")
        
        raw_list = list(soup.find_all('tr'))
        extracted_info = []
        for i in range(0, len(raw_list)):
            if ">Net Conditioned Building Area<" in str(raw_list[i]):
                net_conditioned_area = float(str(raw_list[i])[-17:-11])
            if ">Total Building Area<" in str(raw_list[i]):
                total_area = float(str(raw_list[i])[-17:-11])
        #originally, index -21:-12
        print(f"{cz},conditioned_area = {net_conditioned_area}(m2), total_area = {total_area}(m2)")
        conditioned_area_list_2s.append(net_conditioned_area)
# %%
#create list to be pasted into area field
conditioned_area_1s = []
for i in conditioned_area_list:
    conditioned_area_1s.append([i*10.764])
# %%
#2s area list
conditioned_area_2s = []
for i in conditioned_area_list_2s:
    conditioned_area_2s.append([i*10.764])
# %%
#(1-story)
#Paste in Area (note 1 story or 2 story)
wb_destination2 = wb['Summary']

pasteRange(18,3,18,18, wb_destination2, conditioned_area_1s)
wb.save('Fan_Control_Savings_1_story_v3.xlsx')

# %%
#(2-story)
#Paste in Area (note 1 story or 2 story)
wb_destination2 = wb['Summary']

pasteRange(18,3,18,18, wb_destination2, conditioned_area_2s)
wb.save('Fan_Control_Savings_2_story_v3.xlsx')
# %%
